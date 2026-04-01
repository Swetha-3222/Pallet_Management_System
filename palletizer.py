# palletizer.py
import streamlit as st
import copy
import math
import os
import datetime
import tempfile
import pandas as pd
import plotly.graph_objects as go
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# -------------------- Page --------------------
st.set_page_config(page_title="Palletizer", layout="wide")
st.markdown("<h2 style='text-align:center; color:#008080;'>JK Fenner Palletizer Dashboard</h2>", unsafe_allow_html=True)

# -------------------- Defaults --------------------
DEFAULT_PALLET = {'L': 48.0, 'W': 40.0, 'H': 36.0}
DEFAULT_BOXES = {
    'AZ17': [40.0, 24.0, 9.0],
    'AZ13': [40.0, 16.0, 9.0],
    'AZ6':  [40.0, 11.3, 9.0],
    'AZ16': [24.0, 20.0, 9.0],
    'AZ4':  [24.0, 10.0, 9.0],
    'AZ3':  [24.0, 8.0, 9.0],
    'AZ15': [22.9, 19.3, 9.0],
    'AZ11': [22.7, 13.0, 9.0],
    'AZ14': [22.375, 18.75, 9.0],
    'AZ10': [22.375, 12.75, 9.0],
    'AZ12': [20.0, 16.0, 9.0],
    'AZ8':  [18.375, 11.6, 9.0],
    'AZ5':  [18.375, 11.0, 9.0],
    'AZ7':  [12.0, 11.375, 9.0],
    'AZ2':  [12.0, 8.0, 9.0],
    'AZ18': [48.0, 40.0, 18.0]
}
MOQ = 10  # units per box

# ---------------------------------------------------
# RULE CONFIGURATION
# ---------------------------------------------------

BOX_GROUP_RULES = {
    "AZ3":  ["AZ4"],
    "AZ4":  ["AZ3"],
    "AZ5":  ["AZ2", "AZ6"],
    "AZ10": ["AZ7", "AZ2"],
    "AZ14": ["AZ5", "AZ16"],
}
# ---------------------------------------------------
# MAX BOXES PER LAYER TABLE (NEW LOGIC)
# ---------------------------------------------------

MAX_PER_LAYER = {
    "AZ2": 16,
    "AZ3": 8,
    "AZ4": 8,
    "AZ5": 8,
    "AZ6": 4,
    "AZ7": 12,
    "AZ8": 8,
    "AZ10": 6,
    "AZ11": 6,
    "AZ12": 6,
    "AZ13": 3,
    "AZ14": 4,
    "AZ15": 4,
    "AZ16": 4,
    "AZ17": 2,
    "AZ18": 1
}

FILLER_BOX = "AZ2"
STRICT_SINGLE_PALLET = ["AZ18"]
CONDITIONAL_SEPARATE = ["AZ13", "AZ17"]

def conflicts(box_a, box_b):
    """Return True if box_a and box_b are not allowed together"""
    if box_a in BOX_GROUP_RULES:
        if box_b in BOX_GROUP_RULES[box_a]:
            return True
    if box_b in BOX_GROUP_RULES:
        if box_a in BOX_GROUP_RULES[box_b]:
            return True
    return False


def pallet_has_conflict(pallet_obj, new_box, strict=True):
    if not strict:
        return False

    for layer in pallet_obj['layers']:
        for placed in layer['boxes']:
            if conflicts(placed['name'], new_box):
                return True
    return False

def normalize_box_name(name: str) -> str:
    return ''.join(ch for ch in str(name).upper().strip() if ch.isalnum())

# Normalize DEFAULT_BOXES keys
DEFAULT_BOXES = {normalize_box_name(k): v for k, v in DEFAULT_BOXES.items()}

# -------------------- Sidebar Inputs --------------------
with st.sidebar:
    if os.path.exists(os.path.join(os.path.dirname(__file__), "jk_fenner_logo.png")):
        st.image("jk_fenner_logo.png", width=150)
    st.markdown("### Pallet settings & Inputs")
    pallet_L = st.number_input("Pallet length (in)", value=float(DEFAULT_PALLET['L']))
    pallet_W = st.number_input("Pallet width (in)", value=float(DEFAULT_PALLET['W']))
    pallet_H = st.number_input("Pallet height (in)", value=float(DEFAULT_PALLET['H']))
    
    # ✅ NEW OPTIONAL INPUTS
    st.markdown("#### Optional Customer Details")
    customer_name = st.text_input("Customer Name (alphabets only)")
    invoice_no = st.text_input("Invoice Number (whole numbers)")
    po_ref_no = st.text_input("PO Reference Number (whole numbers)")
    dc = st.text_input("Direct Customer (alphabets only)")

    scale = st.number_input("Scale factor (visual)", value=8, min_value=1)
    st.markdown("---")
    st.markdown("Paste orders **(Part, Qty, Package No, Box No)** — one per line")
    st.markdown("Example: `E71531, 870, 100245, 20`")

    paste_text = st.text_area(
        "Paste Orders (one per line, comma / tab / space separated)",
        height=240
    )
    
    st.markdown("---")
    show_pdf = st.checkbox("Enable PDF download buttons", value=True)
    enable_reserve = st.checkbox("Enable reserve small remainders (legacy behavior)", value=False)

# -------------------- Load mapping --------------------
mapping_file = os.path.join(os.path.dirname(__file__), "MASTER PART.xlsx")
try:
    df_map = pd.read_excel(mapping_file)
    df_map.columns = [c.upper() for c in df_map.columns]
    part_col = [c for c in df_map.columns if "PART" in c.upper()][0]
    box_col = [c for c in df_map.columns if "BOX" in c.upper()][0]
    part_series = df_map[part_col].astype(str).apply(lambda x: x.upper())
    box_series = df_map[box_col].astype(str).apply(normalize_box_name)
    PART_TO_BOX = dict(zip(part_series, box_series))
    BOX_TO_PARTS = {}
    for p, b in PART_TO_BOX.items():
        BOX_TO_PARTS.setdefault(b, []).append(p)
    ALL_PARTS = sorted(PART_TO_BOX.keys())
except Exception as e:
    st.warning(f"MASTER PART.xlsx not found or failed to load: {e}")
    PART_TO_BOX, BOX_TO_PARTS, ALL_PARTS = {}, {}, []

# -------------------- Load AZ WEIGHT MASTER (by PART / SIZE) --------------------
weight_file = os.path.join(os.path.dirname(__file__), "AZ WEIGHT MASTER.xlsx")

try:
    df_wt = pd.read_excel(weight_file)
    df_wt.columns = [c.upper() for c in df_wt.columns]

    # Expecting columns: SIZE, GR WT, NT WT
    if not {"SIZE", "GR WT", "NT WT"}.issubset(df_wt.columns):
        raise ValueError("AZ WEIGHT MASTER.xlsx must contain SIZE, GR WT, NT WT columns")

    WEIGHT_MAP = dict(
        zip(
            df_wt["SIZE"].astype(str).str.upper(),
            zip(df_wt["GR WT"], df_wt["NT WT"])
        )
    )

except Exception as e:
    st.error(f"Failed to load AZ WEIGHT MASTER.xlsx: {e}")
    WEIGHT_MAP = {}



# -------------------- Colors --------------------
if 'colors' not in st.session_state:
    st.session_state.colors = {}
for key in list(ALL_PARTS) + list(DEFAULT_BOXES.keys()):
    if key not in st.session_state.colors:
        h = abs(hash(key)) % (256**3)
        r = (h >> 16) & 0xFF
        g = (h >> 8) & 0xFF
        b = h & 0xFF
        r = 80 + (r % 160)
        g = 80 + (g % 160)
        b = 80 + (b % 160)
        st.session_state.colors[key] = f"rgba({r},{g},{b},0.85)"
st.session_state.colors.setdefault("", "rgba(200,200,200,0.6)")

# -------------------- MaxRects Implementation (2D) --------------------
# Based on common MaxRects variations (best short side fit as default).
class Rect:
    def __init__(self, x, y, w, h, name=None):
        self.x = x
        self.y = y
        self.w = w
        self.h = h
        self.name = name

    def area(self):
        return self.w * self.h

    def __repr__(self):
        return f"Rect(x={self.x},y={self.y},w={self.w},h={self.h},name={self.name})"

class MaxRectsBin:
    def __init__(self, width, height, allow_rotate=True):
        self.width = width
        self.height = height
        self.allow_rotate = allow_rotate
        self.free_rects = [Rect(0, 0, width, height)]
        self.used_rects = []

    def _find_position_for_new_node_best_short_side_fit(self, w, h):
        best_rect = None
        best_short_side = None
        best_long_side = None
        best_x = best_y = 0
        best_rotated = False

        for fr in self.free_rects:
            # try without rotation
            if w <= fr.w + 1e-9 and h <= fr.h + 1e-9:
                leftover_h = abs(fr.h - h)
                leftover_w = abs(fr.w - w)
                short_side = min(leftover_h, leftover_w)
                long_side = max(leftover_h, leftover_w)
                if best_rect is None or (short_side < best_short_side) or (short_side == best_short_side and long_side < best_long_side):
                    best_rect = fr
                    best_short_side = short_side
                    best_long_side = long_side
                    best_x, best_y = fr.x, fr.y
                    best_rotated = False
            # try with rotation if allowed
            if self.allow_rotate and h <= fr.w + 1e-9 and w <= fr.h + 1e-9:
                leftover_h = abs(fr.h - w)
                leftover_w = abs(fr.w - h)
                short_side = min(leftover_h, leftover_w)
                long_side = max(leftover_h, leftover_w)
                if best_rect is None or (short_side < best_short_side) or (short_side == best_short_side and long_side < best_long_side):
                    best_rect = fr
                    best_short_side = short_side
                    best_long_side = long_side
                    best_x, best_y = fr.x, fr.y
                    best_rotated = True
        if best_rect is None:
            return None
        return (best_x, best_y, w if not best_rotated else h, h if not best_rotated else w, best_rotated, best_rect)

    def _split_free_rect(self, free_rect, used):
        new_rects = []
        # used and free rect overlap check
        if used.x >= free_rect.x + free_rect.w or used.x + used.w <= free_rect.x or used.y >= free_rect.y + free_rect.h or used.y + used.h <= free_rect.y:
            # no overlap
            return [free_rect]
        # split horizontally
        if used.x > free_rect.x and used.x < free_rect.x + free_rect.w:
            new_rects.append(Rect(free_rect.x, free_rect.y, used.x - free_rect.x, free_rect.h))
        if used.x + used.w < free_rect.x + free_rect.w:
            new_rects.append(Rect(used.x + used.w, free_rect.y, (free_rect.x + free_rect.w) - (used.x + used.w), free_rect.h))
        # split vertically
        if used.y > free_rect.y and used.y < free_rect.y + free_rect.h:
            new_rects.append(Rect(free_rect.x, free_rect.y, free_rect.w, used.y - free_rect.y))
        if used.y + used.h < free_rect.y + free_rect.h:
            new_rects.append(Rect(free_rect.x, used.y + used.h, free_rect.w, (free_rect.y + free_rect.h) - (used.y + used.h)))
        return new_rects

    def _prune_free_list(self):
        pruned = []
        for i, r in enumerate(self.free_rects):
            overlapped = False
            for j, r2 in enumerate(self.free_rects):
                if i != j and r.x >= r2.x - 1e-9 and r.y >= r2.y - 1e-9 and (r.x + r.w) <= (r2.x + r2.w) + 1e-9 and (r.y + r.h) <= (r2.y + r2.h) + 1e-9:
                    overlapped = True
                    break
            if not overlapped:
                pruned.append(r)
        self.free_rects = pruned

    def insert(self, name, w, h):
        pos = self._find_position_for_new_node_best_short_side_fit(w, h)
        if pos is None:
            return None
        x, y, pw, ph, rotated, used_fr = pos
        used = Rect(x, y, pw, ph, name)
        self.used_rects.append(used)
        # split free rects
        new_free = []
        for fr in self.free_rects:
            splitted = self._split_free_rect(fr, used)
            new_free.extend(splitted)
        self.free_rects = new_free
        self._prune_free_list()
        return {'name': name, 'x': used.x, 'y': used.y, 'L': used.w, 'W': used.h, 'rotated': rotated}

# -------------------- Packer: pack one layer using MaxRects --------------------
def pack_one_layer_maxrects(pallet_L, pallet_W, boxes_info, order_left, current_layer_index, ignore_group_rules=False ):

    bin_pack = MaxRectsBin(pallet_L, pallet_W, allow_rotate=True)
    placed = []

    sequence = build_rule_sequence(order_left)

    for code in sequence:

        while order_left.get(code, 0) > 0:

            # AZ13 & AZ17 not allowed in bottom 2 layers
            if code in CONDITIONAL_SEPARATE and current_layer_index < 2:
                break

            if not ignore_group_rules:
                if not is_group_valid(code, order_left):
                    break


            dims = boxes_info[code]
            result = bin_pack.insert(code, dims['L'], dims['W'])

            if not result:
                break

            placed.append({
                "name": code,
                "x": result['x'],
                "y": result['y'],
                "L": result['L'],
                "W": result['W'],
                "H": dims['H'],
                "rotated": result['rotated']
            })

            order_left[code] -= 1

    return placed, order_left

# -------------------- High-level: pack all pallets by layering --------------------
def build_info(boxes):
    info = {}
    for nm, dims in boxes.items():
        L, W, H = dims
        info[nm] = {'L': float(L), 'W': float(W), 'H': float(H), 'Area': float(L)*float(W)}
    return info

# ===============================
# MAIN PACKING FUNCTION
# ===============================

# ---------------------------------------------------
# PRE-PACK FULL LAYERS BASED ON MAX_PER_LAYER
# ---------------------------------------------------

def prepack_full_layers(order_counts, pallet, boxes):

    pallets = []
    remaining_orders = order_counts.copy()

    pallet_L = pallet["L"]
    pallet_W = pallet["W"]
    pallet_H = pallet["H"]

    for box_code, qty in order_counts.items():

        if box_code not in MAX_PER_LAYER:
            continue

        per_layer = MAX_PER_LAYER[box_code]

        box_L, box_W, box_H = boxes[box_code]

        layers_per_pallet = int(pallet_H // box_H)

        if layers_per_pallet == 0:
            continue

        boxes_per_pallet = per_layer * layers_per_pallet

        # 🚨 ONLY PREPACK FULL PALLETS
        if qty < boxes_per_pallet:
            continue

        pallets_possible = qty // boxes_per_pallet

        for _ in range(pallets_possible):

            pallet_layers = []

            for layer_index in range(layers_per_pallet):

                layer = []

                # Check normal orientation
                rows_normal = int(pallet_W // box_W)
                cols_normal = int(pallet_L // box_L)
                fit_normal = rows_normal * cols_normal

                # Check rotated orientation
                rows_rot = int(pallet_W // box_L)
                cols_rot = int(pallet_L // box_W)
                fit_rot = rows_rot * cols_rot

                # Choose orientation that supports MAX_PER_LAYER
                if fit_rot >= per_layer and fit_rot > fit_normal:

                    rotated = True
                    rows = rows_rot
                    cols = cols_rot
                    place_L = box_W
                    place_W = box_L

                else:

                    rotated = False
                    rows = rows_normal
                    cols = cols_normal
                    place_L = box_L
                    place_W = box_W

                count = 0

                for r in range(rows):
                    for c in range(cols):

                        if count >= per_layer:
                            break

                        x = c * place_L
                        y = r * place_W

                        layer.append({
                            "name": box_code,
                            "x": x,
                            "y": y,
                            "L": place_L,
                            "W": place_W,
                            "H": box_H,
                            "rotated": rotated
                        })

                        count += 1

                    if count >= per_layer:
                        break
                pallet_layers.append(layer)

            pallets.append(pallet_layers)

            remaining_orders[box_code] -= boxes_per_pallet

    return pallets, remaining_orders

def pack_all_pallets_maxrects(pallet, boxes, order_counts):

    info = build_info(boxes)

    # Expand boxes
    all_boxes = []
    for code, qty in order_counts.items():
        for _ in range(qty):
            all_boxes.append(code)

    # Global First-Fit Decreasing
    all_boxes.sort(key=lambda x: info[x]['Area'], reverse=True)

    pallets = []
    unplaced_phase1 = []

    def try_place(box_code, strict=True):

        box_height = info[box_code]['H']

        for pallet_obj in pallets:

            # Conflict check
            if pallet_has_conflict(pallet_obj, box_code, strict):
                continue

            # Try existing layers
            for layer_index, layer_obj in enumerate(pallet_obj['layers']):

                # Stability rule
                if box_code in CONDITIONAL_SEPARATE and layer_index > 1:
                    continue

                result = layer_obj['bin'].insert(
                    box_code,
                    info[box_code]['L'],
                    info[box_code]['W']
                )

                if result:
                    layer_obj['boxes'].append({
                        "name": box_code,
                        "x": result['x'],
                        "y": result['y'],
                        "L": result['L'],
                        "W": result['W'],
                        "H": box_height,
                        "rotated": result['rotated']
                    })
                    return True

            # Try new layer
            current_height = sum(l['height'] for l in pallet_obj['layers'])

            if current_height + box_height <= pallet['H']:

                new_layer_index = len(pallet_obj['layers'])

                if box_code in CONDITIONAL_SEPARATE and new_layer_index > 1:
                    continue

                new_bin = MaxRectsBin(pallet['L'], pallet['W'], allow_rotate=True)

                result = new_bin.insert(
                    box_code,
                    info[box_code]['L'],
                    info[box_code]['W']
                )

                if result:
                    pallet_obj['layers'].append({
                        'bin': new_bin,
                        'boxes': [{
                            "name": box_code,
                            "x": result['x'],
                            "y": result['y'],
                            "L": result['L'],
                            "W": result['W'],
                            "H": box_height,
                            "rotated": result['rotated']
                        }],
                        'height': box_height
                    })
                    return True

        return False

    # -------------------------
    # PHASE 1 — STRICT
    # -------------------------
    for box_code in all_boxes:

        placed = try_place(box_code, strict=True)

        if not placed:
            unplaced_phase1.append(box_code)

    # -------------------------
    # PHASE 2 — RELAX RULES
    # -------------------------
    for box_code in unplaced_phase1:

        placed = try_place(box_code, strict=False)

        if not placed:
            # Create new pallet only if absolutely necessary
            new_bin = MaxRectsBin(pallet['L'], pallet['W'], allow_rotate=True)

            result = new_bin.insert(
                box_code,
                info[box_code]['L'],
                info[box_code]['W']
            )

            pallets.append({
                'layers': [{
                    'bin': new_bin,
                    'boxes': [{
                        "name": box_code,
                        "x": result['x'],
                        "y": result['y'],
                        "L": result['L'],
                        "W": result['W'],
                        "H": info[box_code]['H'],
                        "rotated": result['rotated']
                    }],
                    'height': info[box_code]['H']
                }]
            })

    pallet_layers = []
    for p in pallets:
        pallet_layers.append([layer['boxes'] for layer in p['layers']])

    print(f"\n✅ Total pallets used: {len(pallet_layers)}")

    return pallet_layers

def is_group_valid(code, order_left):
    if code in BOX_GROUP_RULES:
        for partner in BOX_GROUP_RULES[code]:
            if order_left.get(partner, 0) <= 0:
                return False
    return True

def build_rule_sequence(order_left):
    sequence = []
    visited = set()

    for code in order_left:
        if order_left[code] <= 0 or code in visited:
            continue

        if code in BOX_GROUP_RULES:
            group = [code] + BOX_GROUP_RULES[code]
            for g in group:
                if order_left.get(g, 0) > 0:
                    sequence.append(g)
                    visited.add(g)
        else:
            sequence.append(code)
            visited.add(code)

    # AZ2 as filler → always last
    if FILLER_BOX in sequence:
        sequence.remove(FILLER_BOX)
        sequence.append(FILLER_BOX)

    return sequence


# -------------------- Parse pasted orders --------------------
order_counts = {}
order_part_queue = {}
box_pkg_queue = {}   # box_code → list of package numbers (1 per box)
box_box_queue = {}   # box_code → list of box numbers (1 per box)

for line in paste_text.splitlines():
    s = line.strip()
    if not s:
        continue

    # allow comma / tab / space separation
    tokens = [t.strip() for t in s.replace("\t", ",").split(",") if t.strip()]

    if len(tokens) < 2:
        st.warning(f"Invalid line skipped: {line}")
        continue

    part = tokens[0]
    qty = tokens[1]
    

    # Package number (optional, ignored for planning)
    pkg_no = None
    box_no = None
    if len(tokens) >= 3:
        try:
            pkg_no = int(tokens[2])  # stored if needed later
        except:
            st.warning(f"Invalid package number ignored in line: {line}")

    if len(tokens) >= 4:
        try:
            box_no = int(tokens[3])  # stored if needed later
        except:
            st.warning(f"Invalid box number ignored in line: {line}")

    part_up = part.upper()

    if part_up in DEFAULT_BOXES:
        st.warning(
            f"Detected box code '{part_up}' in pasted input. "
            f"Please paste PART numbers only. This line was skipped."
        )
        continue

    if part_up not in PART_TO_BOX:
        st.warning(f"Unknown part '{part_up}' skipped.")
        continue

    try:
        qty_units = int(float(qty))
    except:
        st.warning(f"Bad qty on line '{line}' — skipped.")
        continue

    boxes_needed = math.ceil(qty_units / MOQ)
    box_code = normalize_box_name(PART_TO_BOX[part_up])

    order_counts[box_code] = order_counts.get(box_code, 0) + boxes_needed
    order_part_queue.setdefault(box_code, []).extend([part_up] * boxes_needed)

    # NEW: generate unique package numbers per box
    if pkg_no is not None:
        box_pkg_queue.setdefault(box_code, [])
        for i in range(boxes_needed):
            box_pkg_queue[box_code].append(pkg_no + i)
    # NEW: generate unique package numbers per box
    if box_no is not None:
        box_box_queue.setdefault(box_code, [])
        for i in range(boxes_needed):
            box_box_queue[box_code].append(box_no + i)

# ensure default boxes keys exist
for b in DEFAULT_BOXES.keys():
    order_counts.setdefault(b, 0)
    order_part_queue.setdefault(b, [])

if not any(v > 0 for v in order_counts.values()):
    st.info("No valid parts found from pasted orders. Paste PART numbers & quantities to pack.")
    st.stop()


# -------------------- Prepare boxes_for_packing (allow rotation automatically handled by packer) --------------------
boxes_for_packing = copy.deepcopy(DEFAULT_BOXES)

# -------------------- Run packer --------------------
pallet = {'L': float(pallet_L), 'W': float(pallet_W), 'H': float(pallet_H)}
reserve_flag = enable_reserve
# -------------------- Run packing --------------------

# STEP 1: Pre-pack full pallets with same box type
pre_pallets, remaining_orders = prepack_full_layers(
    order_counts,
    pallet,
    boxes_for_packing
)

# STEP 2: Run optimizer on remaining boxes
optimized_pallets = pack_all_pallets_maxrects(
    pallet,
    boxes_for_packing,
    remaining_orders
)

# STEP 3: Combine pallets
pallet_layers = pre_pallets + optimized_pallets
st.success(f"Total pallets used: {len(pallet_layers)}")
total_pallets = len(pallet_layers)

parts = set()

for pal in pallet_layers:
    for layer in pal:
        for box in layer:
            part = box.get("part")
            if part:
                parts.add(part)

if len(parts) == 1:
    pass

    # do not finalize pallet
    # push back boxes to order_left


# -------------------- Assign parts to placed boxes for summary & visuals --------------------
local_queues_for_summary = {k: v.copy() for k, v in order_part_queue.items()}
summary_per_pallet = []
grand_total = {}

assigned_layers_per_pallet = []
for pal_layers in pallet_layers:
    assigned_layers_per_pallet.append([])
    pal_layer_dicts = []
    pal_totals = {}
    for layer in pal_layers:
        layer_counts = {}
        # The layer is list of boxes with 'name' code; assign part from queue if available
        for b in layer:
            box_code = b['name']
            qlist = local_queues_for_summary.get(box_code, [])
            if qlist:
                part_assigned = qlist.pop(0)
            else:
                part_assigned = BOX_TO_PARTS.get(box_code, [''])[0] or ''
            b['part'] = part_assigned
            key = f"{part_assigned} ({box_code})" if part_assigned else box_code
            layer_counts[key] = layer_counts.get(key, 0) + 1
            pal_totals[key] = pal_totals.get(key, 0) + 1
            grand_total[key] = grand_total.get(key, 0) + 1
        pal_layer_dicts.append(layer_counts)
        assigned_layers_per_pallet[-1].append(layer)
    summary_obj = {'_layers': pal_layer_dicts, '_pallet_totals': {'total_boxes': sum(pal_totals.values())}}
    for k, v in pal_totals.items():
        summary_obj[k] = v
    summary_per_pallet.append(summary_obj)

# -------------------- Visual: top view with Plotly --------------------
def plot_layer_topview(pallet, layer_data, colors_map, scale=8, title=None, show_labels=True):
    L_px = pallet['L'] * scale
    W_px = pallet['W'] * scale
    fig = go.Figure()
    fig.add_shape(type='rect', x0=0, y0=0, x1=L_px, y1=W_px, line=dict(width=2, color='black'), fillcolor="rgba(0,0,0,0)")
    for d in layer_data:
        x0 = d['x'] * scale
        y0 = d['y'] * scale
        x1 = x0 + d['L'] * scale
        y1 = y0 + d['W'] * scale
        part = d.get('part','')
        color_key = part if part else d['name']
        color = colors_map.get(color_key, colors_map.get(d['name'], "rgba(200,200,200,0.7)"))
        fig.add_shape(type='rect', x0=x0, y0=y0, x1=x1, y1=y1,
                      line=dict(color='black', width=1), fillcolor=color)
        if show_labels:
            label = f"{part} ({d['name']})" if part else d['name']
            fig.add_annotation(x=(x0+x1)/2, y=(y0+y1)/2, text=label, showarrow=False,
                               font=dict(size=9), xanchor='center', yanchor='middle', font_color="white")
    fig.update_xaxes(showticklabels=False, range=[0, L_px])
    fig.update_yaxes(showticklabels=False, range=[0, W_px], scaleanchor='x', scaleratio=1)
    fig.update_layout(width=600, height=int(600 * (pallet['W'] / pallet['L'] + 0.02)), margin=dict(l=10, r=10, t=30, b=10))
    if title:
        fig.update_layout(title=title)
    return fig

# -------------------- PDF generation --------------------
def ddmmyyyy_hhmmss_now():
    return datetime.datetime.now().strftime("%d%m%Y_%H%M%S")

def create_summary_pdf_text(summary_per_pallet, grand_total, pallet):
    ts = ddmmyyyy_hhmmss_now()
    path = tempfile.NamedTemporaryFile(delete=False, suffix=f"_summary_{ts}.pdf").name
    c = canvas.Canvas(path, pagesize=A4)
    pw, ph = A4
    margin = 36
    y = ph - 60
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Pallet Packing - Summary Report")
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(margin, y, f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
    y -= 18
    total_pallets = len(summary_per_pallet)
    c.drawString(margin, y, f"Total Pallets: {total_pallets}")
    y -= 20

    for p_idx, pal_sum in enumerate(summary_per_pallet, start=1):
        if y < 120:
            c.showPage(); y = ph - 60
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, f"Pallet {p_idx}")
        y -= 14
        c.setFont("Helvetica", 10)
        layers = pal_sum.get('_layers', [])
        if layers:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(margin+6, y, "Layer")
            c.drawString(margin+120, y, "Box Type")
            c.drawString(margin+260, y, "Boxes")
            y -= 12
            c.setFont("Helvetica", 10)
            for lnum, layer_dict in enumerate(layers, start=1):
                for boxk, cnt in layer_dict.items():
                    if y < 80:
                        c.showPage(); y = ph - 60
                    c.drawString(margin+6, y, f"{lnum}")
                    c.drawString(margin+120, y, f"{boxk}")
                    c.drawString(margin+260, y, f"{cnt}")
                    y -= 12
                y -= 6
        pal_tot = pal_sum.get('_pallet_totals', {})
        if y < 80:
            c.showPage(); y = ph - 60
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin+6, y, "Pallet Totals:")
        y -= 12
        c.setFont("Helvetica", 10)
        total_boxes = pal_tot.get('total_boxes', sum([v for k,v in pal_sum.items() if not k.startswith('_')]))
        total_material = total_boxes * MOQ
        c.drawString(margin+20, y, f"Total boxes: {total_boxes}")
        y -= 12
        c.drawString(margin+20, y, f"Total material (units): {total_material}")
        y -= 18

    if y < 120:
        c.showPage(); y = ph - 60
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, y, "Grand Totals")
    y -= 14
    c.setFont("Helvetica", 10)
    total_boxes_all = sum(v for k,v in grand_total.items())
    total_material_all = total_boxes_all * MOQ
    c.drawString(margin+10, y, f"Total boxes across all pallets: {total_boxes_all}")
    y -= 12
    c.drawString(margin+10, y, f"Total material across all pallets (units): {total_material_all}")
    y -= 12

    c.save()
    return path

def create_layout_pdf_visuals(assigned_layer_details, pallet):
    ts = ddmmyyyy_hhmmss_now()
    path = tempfile.NamedTemporaryFile(delete=False, suffix=f"_layout_{ts}.pdf").name
    c = canvas.Canvas(path, pagesize=A4)
    pw, ph = A4
    margin = 20
    cols = 2
    rows = 4
    thumb_w = (pw - margin*2 - (cols-1)*6) / cols
    thumb_h = (ph - margin*2 - 60 - (rows-1)*6) / rows

    for p_idx, pal_layers in enumerate(assigned_layer_details, start=1):
        i = 0
        while i < len(pal_layers):
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, ph - margin - 10, f"Pallet {p_idx} - Layout (Generated {datetime.datetime.now().strftime('%d-%b-%Y %H:%M:%S')})")
            for r in range(rows):
                for co in range(cols):
                    idx = i
                    x = margin + co*(thumb_w+6)
                    y_top = ph - margin - 40 - r*(thumb_h+6)
                    c.rect(x, y_top - thumb_h, thumb_w, thumb_h, stroke=1, fill=0)
                    if idx < len(pal_layers):
                        layer = pal_layers[idx]
                        if layer:
                            min_x = min(b['x'] for b in layer)
                            min_y = min(b['y'] for b in layer)
                            max_x = max(b['x'] + b['L'] for b in layer)
                            max_y = max(b['y'] + b['W'] for b in layer)
                            lw = max_x - min_x
                            lh = max_y - min_y
                            if lw <= 0: lw = 1.0
                            if lh <= 0: lh = 1.0
                            sx = (thumb_w - 8) / lw
                            sy = (thumb_h - 8) / lh
                            s = min(sx, sy)
                        else:
                            s = 1.0
                            min_x = min_y = 0.0
                        for b in layer:
                            bx = x + 4 + (b['x'] - min_x) * s
                            by_top = y_top - 4 - (b['y'] - min_y) * s
                            bw = b['L'] * s
                            bh = b['W'] * s
                            c.setFillColorRGB(0.85, 0.9, 1)
                            c.rect(bx, by_top - bh, bw, bh, stroke=1, fill=1)
                            lab = (b.get('part','') + ' (' + b['name'] + ')') if b.get('part','') else b['name']
                            c.setFont('Helvetica', 6)
                            c.drawString(bx+2, by_top - bh/2 - 3, lab[:40])
                        c.setFont('Helvetica-Bold', 9)
                        c.drawString(x+2, y_top - thumb_h - 10, f"Layer {idx+1}")
                    i += 1
                    if i >= len(pal_layers):
                        break
            c.showPage()
    c.save()
    return path

def clear_values_after_row(ws, start_row=7):
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

def create_excel_report(
    assigned_layers_per_pallet,
    customer_name,
    invoice_no,
    po_ref_no,
    dc
):
    template_path = os.path.join(os.path.dirname(__file__), "Report format.xlsx")
    wb = load_workbook(template_path)

    # First sheet is the template
    template_ws = wb.worksheets[0]
    template_ws.title = "Pallet 1"

    global_box_no = 1  # BOX NO must continue across pallets

    def write_pallet(ws, pallet_no, pallet_layers):
        nonlocal global_box_no

        # ---------------- Header ----------------
        ws["F3"] = f"CUSTOMER NAME : {customer_name}"
        ws["F4"] = f"INVOICE NO : {invoice_no}"
        ws["F5"] = f"Pallet {pallet_no}"

        start_row = 7
        current_row = start_row

        # ---------------- Table rows ----------------
        for layer in pallet_layers:
            for b in layer:
                part = b.get("part", "").upper()

                # Get PKG NO safely
                box_code = b["name"]
                pkg_no = ""
                box_no = ""
                if box_code in box_pkg_queue and box_pkg_queue[box_code]:
                    pkg_no = box_pkg_queue[box_code].pop(0)
                if box_code in box_box_queue and box_box_queue[box_code]:
                    box_no = box_box_queue[box_code].pop(0)

                gr, nt = WEIGHT_MAP.get(part, ("", ""))

                ws.cell(row=current_row, column=6, value=box_no)  # BOX NO
                ws.cell(row=current_row, column=7, value=pkg_no)         # PKGS NO
                ws.cell(row=current_row, column=8, value=part)           # SIZE
                ws.cell(row=current_row, column=9, value=MOQ)            # QTY
                ws.cell(row=current_row, column=10, value=gr)            # GR WT
                ws.cell(row=current_row, column=11, value=nt)            # NT WT
                ws.cell(row=current_row, column=12, value=po_ref_no)    
                ws.cell(row=current_row, column=13, value=dc)            # Direct Customer

                global_box_no += 1
                current_row += 1

        # ---------------- TOTAL row ----------------
        total_row = current_row
        total_pkgs = total_row - start_row

        ws.insert_rows(total_row)

        ws.merge_cells(start_row=total_row, start_column=6, end_row=total_row, end_column=8)
        ws.cell(row=total_row, column=6, value=f"{total_pkgs} PKGS")

        ws.cell(row=total_row, column=9, value=total_pkgs * MOQ)
        ws.cell(row=total_row, column=10, value=f"=SUM(J{start_row}:J{total_row-1})")
        ws.cell(row=total_row, column=11, value=f"=SUM(K{start_row}:K{total_row-1})")
        ws.cell(row=total_row, column=12, value="")

    # ---------------- Write Pallet 1 ----------------
    write_pallet(template_ws, 1, assigned_layers_per_pallet[0])

    # ---------------- Copy template for remaining pallets ----------------
    for p_idx in range(2, len(assigned_layers_per_pallet) + 1):
        ws_new = wb.copy_worksheet(template_ws)
        ws_new.title = f"Pallet {p_idx}"

        # ---- Unmerge cells from row 7 onwards ----
        merged_ranges = list(ws_new.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.min_row >= 7:
                ws_new.unmerge_cells(str(merged_range))

        clear_values_after_row(ws_new, start_row=7)

        write_pallet(ws_new, p_idx, assigned_layers_per_pallet[p_idx - 1])


    # ---------------- Save output ----------------
    out_path = tempfile.NamedTemporaryFile(
        delete=False,
        suffix="_Invoice_Report.xlsx"
    ).name

    wb.save(out_path)
    return out_path

# -------------------- Top bar & PDF buttons --------------------
st.markdown("---")
top_cols = st.columns([2, 4, 1.2, 1.2])
with top_cols[0]:
    st.markdown(f"<div style='display:flex; align-items:center; gap:8px;'><h4 style='margin:0;'> Total pallets used: {total_pallets}</h4></div>", unsafe_allow_html=True)
with top_cols[1]:
    st.write("")
with top_cols[2]:
    try:
        report_path = create_excel_report(
            assigned_layers_per_pallet,
            customer_name,
            invoice_no,
            po_ref_no,
            dc
        )
        with open(report_path, "rb") as f:
            st.download_button(
                "📊 Download Report",
                f,
                file_name=os.path.basename(report_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Report generation failed: {e}")
with top_cols[3]:
    if show_pdf:
        try:
            summary_pdf_path = create_summary_pdf_text(summary_per_pallet, grand_total, pallet)
            with open(summary_pdf_path, 'rb') as f:
                st.download_button("📝 Download Summary PDF", f, file_name=os.path.basename(summary_pdf_path), mime="application/pdf")
        except Exception as e:
            st.error(f"Error preparing Summary PDF: {e}")


st.markdown("---")

# -------------------- Show Pallets & Layers --------------------
local_queues = {k: v.copy() for k, v in order_part_queue.items()}
for p_idx, layers in enumerate(pallet_layers, start=1):
    st.markdown(f"## 🟫 Pallet {p_idx}")
    for l_idx, layer in enumerate(layers, start=1):
        fig = plot_layer_topview(pallet, layer, st.session_state.colors, scale=scale, title=f"Pallet {p_idx} — Layer {l_idx}", show_labels=True)
        st.plotly_chart(fig, key=f"final_view_p{p_idx}_l{l_idx}", use_container_width=True)

st.caption("Use the Download buttons above to get the Layout and Summary PDFs.")
